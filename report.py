from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
REPORT_DIR = BASE_DIR / "reports"


HEADERS = [
    "Run time",
    "Change",
    "Status",
    "Title",
    "Role",
    "URL",
    "Source",
    "HR / contact emails",
    "Application links",
    "Details",
    "Applied?"
]


def get_applied_urls(base_dir=None):
    """Return all job URLs manually marked as Applied in the Excel report."""
    base = Path(base_dir) if base_dir else BASE_DIR

    candidates = [
        base / "job_report.xlsx",
        base / "reports" / "job_report.xlsx",
    ]

    if base.name == "reports":
        candidates = [base / "job_report.xlsx", *candidates]

    report_path = next((path for path in candidates if path.exists()), None)

    if report_path is None:
        return set()

    try:
        from openpyxl import load_workbook

        workbook = load_workbook(
            report_path,
            read_only=True,
            data_only=True
        )
        try:
            sheet = workbook.active

            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return set()

            headers = [str(value).strip().lower() if value is not None else "" for value in rows[0]]
            url_index = headers.index("url") if "url" in headers else None
            applied_index = None

            for candidate in ("applied?", "applied"):
                if candidate in headers:
                    applied_index = headers.index(candidate)
                    break

            if url_index is None or applied_index is None:
                return set()

            applied_urls = set()

            for row in rows[1:]:
                if len(row) <= max(url_index, applied_index):
                    continue

                url_value = row[url_index]
                applied_value = row[applied_index]

                if url_value is None:
                    continue

                normalized_url = str(url_value).strip()
                if not normalized_url:
                    continue

                if isinstance(applied_value, str) and applied_value.strip().lower() == "applied":
                    applied_urls.add(normalized_url)

            return applied_urls
        finally:
            workbook.close()

    except Exception:
        return set()


def write_report(rows, base_dir=None):
    base = Path(base_dir) if base_dir else BASE_DIR
    report_dir = base / "reports"

    report_dir.mkdir(exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Job opportunities"
    sheet.append(HEADERS)

    header_fill = PatternFill(
        "solid",
        fgColor="1F4E78"
    )

    for cell in sheet[1]:
        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )
        cell.fill = header_fill

    for row in rows:
        sheet.append([
            row.get("run_time", ""),
            row.get("change", ""),
            row.get("status", ""),
            row.get("title", ""),
            row.get("role", ""),
            row.get("url", ""),
            row.get("source", ""),
            row.get("emails", ""),
            row.get("application_links", ""),
            row.get("details", ""),
            row.get("applied_status", "")
        ])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    widths = [
        22,
        16,
        18,
        42,
        24,
        70,
        24,
        36,
        70,
        52,
        18
    ]

    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[
            get_column_letter(index)
        ].width = width

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = cell.alignment.copy(
                vertical="top",
                wrap_text=True
            )

    current_path = report_dir / "job_report.xlsx"

    workbook.save(current_path)

    return current_path
